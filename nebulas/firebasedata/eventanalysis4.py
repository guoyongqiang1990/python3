# -*- coding: utf-8 -*-
import os, sys
import csv
import openpyxl
import xlwt

#读取从firebase导出的csv埋点数据

def read_csvdata(csvpath):
    try:
        with open(csvpath, 'r') as csvfile:
            reader = csv.reader(csvfile)
            csvdata = [row for row in reader]
        return csvdata[4:]
    except IOError as err:
        print("File Error: " + str(err))

# 读取需求excel文档
def read_requirement(requirementpath):
    try:
        with open(requirementpath, 'rb') as requirementfile:
            workbook = openpyxl.load_workbook(requirementfile)
            worksheet = workbook.active
            # print(worksheet.title)
            # 将excel数据按行读取成列表
            requirementdata = []
            for i in range(len(list(worksheet.rows))):
                rowdata = []
                for cell in list(worksheet.rows)[i]:
                    rowdata.append(cell.value)
                if type(rowdata[0]) is int or type(rowdata[0]) is float:  # 将不是埋点事件的行去掉
                    requirementdata.append(rowdata[1:4])
        print(requirementdata)
        return requirementdata
    except IOError as err:
        print("File Error: " + str(err))

#处理数据
def data_analysis(csvdata, requirementdata):
    csvdict = {}
    for items in csvdata:
        csvdict.update({items[0]: items})
    #print(csvdict)
    # 验证埋点事件
    for items in requirementdata:
        if items[1] in csvdict and csvdict[items[1]][1] != "0" and csvdict[items[1]][2] != '0':
            items.append("Pass")
        if items[1] not in csvdict or csvdict[items[1]][1] == '0' or csvdict[items[1]][2] == '0':
            items.append("Fail")

    return requirementdata

'''
#验证埋点需求完整性
def verify_requirement(csvtestresult, requirementdata):
    csvtestresult2 = "".join(str(csvtestresult))
    failedresult = []
    for i in range(len(requirementdata)):
        failedevent=[]
        if requirementdata[i][2] not in csvtestresult2:
            failedevent.extend([requirementdata[i][1], requirementdata[i][2],"统计失败！"])
            #print("%s  %s 事件统计失败!\n" % (requirementdata[i][1], requirementdata[i][2]))
            #print(failedevent)
            failedresult.append(failedevent)
    print(failedresult)
    print("埋点生效情况验证完毕！")
    return failedresult'''

'''
#写数据 保存为csv文件
def write_csvresult(csvdata, csvresultpath):
    try:
        with open(csvresultpath, 'w') as csvfile2:
            writer = csv.writer(csvfile2)
            for eachline in csvdata:
                writer.writerow(eachline)
    except IOError as err:
        print("File Error: " + str(err))
    print("埋点生效情况验证完毕！")


#写数据 保存为excel文件，使用openpyxl
def write_exlresult(csvdata, exlresultpath):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "EventTestresult"
        for i in range(0, len(csvdata)):
            for j in range(0, len(csvdata[i])):
                sheet.cell(row=i+1, column=j+1, value= csvdata[i][j])
        wb.save(exlresultpath)
        print("埋点生效情况验证完毕！")
    except IOError as err:
        print("File Error: " + str(err))'''

#写数据 保存为excel文件，使用xlwt
def write_exlresult(testresult, exlresultpath):
    try:
        wb = xlwt.Workbook()
        sheet = wb.add_sheet("EventTestResult")
        #设置第一列的列宽
        sheet.col(0).width = 256 * 35
        sheet.col(1).width = 256 * 35

        # 初始化样式
        style1 = xlwt.XFStyle()
        # 为样式创建字体,指定名字，加粗
        font1 = xlwt.Font()
        font1.name = 'Times New Roman'
        font1.bold = True
        # 讲该字体设定为style的字体
        style1.font = font1

        style2 = xlwt.XFStyle()
        font2 = xlwt.Font()
        font2.name = "Times New Roman"
        font2.bold = True
        font2.colour_index = 2
        style2.font = font2

        #写数据
        sheet.write(0, 0, "NasNano埋点统计测试结果", style1)
        sheet.write(1, 0, "埋点", style1)
        sheet.write(1, 1, "key", style1)
        sheet.write(1, 2, "版本", style1)
        sheet.write(1, 3, "测试结果", style1)
        for i in range(0, len(testresult)):
            for j in range(0, len(testresult[i])):
                if testresult[i][j] == "Fail":
                    sheet.write(i+2, j, testresult[i][j], style2)
                else:
                    sheet.write(i+2, j, testresult[i][j], style1)
        wb.save(exlresultpath)

    except IOError as err:
        print("File Error: " + str(err))


if __name__ == "__main__":
    requirementpath = os.getcwd() + os.sep + "data" + os.sep + "2.2埋点需求.xlsx"
    filepath = os.getcwd() + os.sep + "data" + os.sep + "data-export.csv"
    #csvresultpath = 'csvtestresult.csv'
    exlresultpath = os.getcwd() + os.sep + "data" + os.sep + "exltestresult.xls"
    csvdata = read_csvdata(filepath)
    requirementdata = read_requirement(requirementpath)
    testresult = data_analysis(csvdata, requirementdata)
    # print(requirementdata)
    #failedresult = verify_requirement(csvtestresult, requirementdata)
    #write_csvresult(finalresult, csvresultpath)
    write_exlresult(testresult, exlresultpath)