#work to do:干掉converse一列，对比key值

# -*- coding: utf-8 -*-
import os, sys
import csv
import openpyxl
import xlwt

#读取csv数据

def read_csvdata(csvpath):
    try:
        with open(csvpath, 'r') as csvfile:
            reader = csv.reader(csvfile)
            csvdata = [row for row in reader]
        return csvdata
    except IOError as err:
        print("File Error: " + str(err))

#处理数据
def data_analysis(csvdata):
    # 增加一列
    csvdata[3].append("result")
    #验证埋点事件
    for i in range(4, len(csvdata)):
        if csvdata[i][1] == "0" or csvdata[i][2] == "0":
            csvdata[i].append('Fail')
        else:
            csvdata[i].append("Pass")
        print(csvdata[i])
    return csvdata
'''
#写数据 保存为csv文件
def write_csvresult(csvdata, csvresultpath):
    try:
        with open(csvresultpath, 'w') as csvfile2:
            writer = csv.writer(csvfile2)
            for eachline in csvdata:
                writer.writerow(eachline)
    except IOError as err:
        print("File Error: " + str(err))
    print("埋点生效情况验证完毕！")


#写数据 保存为excel文件，使用openpyxl
def write_exlresult(csvdata, exlresultpath):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "EventTestresult"
        for i in range(0, len(csvdata)):
            for j in range(0, len(csvdata[i])):
                sheet.cell(row=i+1, column=j+1, value= csvdata[i][j])
        wb.save(exlresultpath)
        print("埋点生效情况验证完毕！")
    except IOError as err:
        print("File Error: " + str(err))'''

#写数据 保存为excel文件，使用xlwt
def write_exlresult(csvdata, exlresultpath):
    try:
        wb = xlwt.Workbook()
        sheet = wb.add_sheet("EventTestResult")
        #设置第一列的列宽
        first_col = sheet.col(0)
        first_col.width = 256*35

        # 初始化样式
        style1 = xlwt.XFStyle()
        # 为样式创建字体,指定名字，加粗
        font1 = xlwt.Font()
        font1.name = 'Times New Roman'
        font1.bold = True
        # 讲该字体设定为style的字体
        style1.font = font1

        style2 = xlwt.XFStyle()
        font2 = xlwt.Font()
        font2.name = "Times New Roman"
        font2.bold = True
        font2.colour_index = 2
        style2.font = font2

        #写数据
        for i in range(0, len(csvdata)):
            for j in range(0, len(csvdata[i])):
                if csvdata[i][j] == "Fail":
                    sheet.write(i, j, csvdata[i][j], style2)
                else:
                    sheet.write(i, j, csvdata[i][j], style1)



        wb.save(exlresultpath)
        print("埋点生效情况验证完毕！")

    except IOError as err:
        print("File Error: " + str(err))

if __name__ == "__main__":
    from sys import argv
    #filepath = argv[1]
    #testresultpath = argv[2]
    filepath = "data-export.csv"
    #csvresultpath = 'csvtestresult.csv'
    exlresultpath = 'exltestresult.xls'
    data = read_csvdata(filepath)
    data2 = data_analysis(data)
    #write_csvresult(data2, csvresultpath)
    write_exlresult(data2, exlresultpath)