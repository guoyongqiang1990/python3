# -*- coding: utf-8 -*-
import os, sys
import csv
import openpyxl
import xlwt

#读取从firebase导出的csv埋点数据

def read_csvdata(csvpath):
    try:
        with open(csvpath, 'r') as csvfile:
            reader = csv.reader(csvfile)
            csvdata = [row for row in reader]
        return csvdata
    except IOError as err:
        print("File Error: " + str(err))

#处理数据
def data_analysis(csvdata):
    # 增加一列
    csvdata[3].append("result")
    # 验证埋点事件
    for i in range(4, len(csvdata)):
        if csvdata[i][1] == "0" or csvdata[i][2] == "0":
            csvdata[i].append('Fail')
        else:
            csvdata[i].append("Pass")
    #将第三列conversion去掉
    for i in range(3, len(csvdata)):
        del csvdata[i][3]
        print(csvdata[i])
    return csvdata

#读取需求excel文档
def read_requirement(requirementpath):
    try:
        with open(requirementpath, 'rb') as requirementfile:
            workbook = openpyxl.load_workbook(requirementfile)
            worksheet = workbook.active
            #print(worksheet.title)
            #将excel数据按行读取成列表
            requirementdata = []
            for i in range(len(list(worksheet.rows))):
                rowdata = []
                for cell in list(worksheet.rows)[i]:
                    rowdata.append(cell.value)
                if type(rowdata[0]) is int or type(rowdata[0]) is float: #将不是埋点事件的行去掉
                    requirementdata.append(rowdata)
        #print(requirementdata)
        return requirementdata
    except IOError as err:
        print("File Error: " + str(err))

#验证埋点需求完整性
def verify_requirement(csvtestresult, requirementdata):
    csvtestresult2 = "".join(str(csvtestresult))
    failedresult = []
    for i in range(len(requirementdata)):
        failedevent=[]
        if requirementdata[i][2] not in csvtestresult2:
            failedevent.extend([requirementdata[i][1], requirementdata[i][2],"统计失败！"])
            #print("%s  %s 事件统计失败!\n" % (requirementdata[i][1], requirementdata[i][2]))
            #print(failedevent)
            failedresult.append(failedevent)
    print(failedresult)
    print("埋点生效情况验证完毕！")
    return failedresult

'''
#写数据 保存为csv文件
def write_csvresult(csvdata, csvresultpath):
    try:
        with open(csvresultpath, 'w') as csvfile2:
            writer = csv.writer(csvfile2)
            for eachline in csvdata:
                writer.writerow(eachline)
    except IOError as err:
        print("File Error: " + str(err))
    print("埋点生效情况验证完毕！")


#写数据 保存为excel文件，使用openpyxl
def write_exlresult(csvdata, exlresultpath):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "EventTestresult"
        for i in range(0, len(csvdata)):
            for j in range(0, len(csvdata[i])):
                sheet.cell(row=i+1, column=j+1, value= csvdata[i][j])
        wb.save(exlresultpath)
        print("埋点生效情况验证完毕！")
    except IOError as err:
        print("File Error: " + str(err))'''

#写数据 保存为excel文件，使用xlwt
def write_exlresult(failedtestresult, csvdata, exlresultpath):
    try:
        wb = xlwt.Workbook()
        sheet = wb.add_sheet("EventTestResult")
        #设置第一列的列宽
        sheet.col(0).width = 256 * 35
        sheet.col(1).width = 256 * 35

        # 初始化样式
        style1 = xlwt.XFStyle()
        # 为样式创建字体,指定名字，加粗
        font1 = xlwt.Font()
        font1.name = 'Times New Roman'
        font1.bold = True
        # 讲该字体设定为style的字体
        style1.font = font1

        style2 = xlwt.XFStyle()
        font2 = xlwt.Font()
        font2.name = "Times New Roman"
        font2.bold = True
        font2.colour_index = 2
        style2.font = font2

        #写数据
        for i in range(0, len(failedtestresult)):
            for j in range(0, len(failedtestresult[i])):
                    sheet.write(i, j, failedtestresult[i][j], style2)

        for i in range(0, len(csvdata)):
            for j in range(0, len(csvdata[i])):
                if csvdata[i][j] == "Fail":
                    sheet.write(i+len(failedtestresult), j, csvdata[i][j], style2)
                else:
                    sheet.write(i+len(failedtestresult), j, csvdata[i][j], style1)

        wb.save(exlresultpath)

    except IOError as err:
        print("File Error: " + str(err))


if __name__ == "__main__":
    from sys import argv
    #requirementpath = argv[1]
    #filepath = argv[2]
    #testresultpath = argv[3]
    requirementpath = os.getcwd() + os.sep + "data" + os.sep + "2.2埋点需求.xlsx"
    filepath = os.getcwd() + os.sep + "data" + os.sep + "data-export.csv"
    #csvresultpath = 'csvtestresult.csv'
    exlresultpath = os.getcwd() + os.sep + "data" + os.sep + "exltestresult.xls"
    csvdata = read_csvdata(filepath)
    csvtestresult = data_analysis(csvdata)
    requirementdata = read_requirement(requirementpath)
    # print(requirementdata)
    failedresult = verify_requirement(csvtestresult, requirementdata)
    #write_csvresult(finalresult, csvresultpath)
    write_exlresult(failedresult, csvtestresult, exlresultpath)